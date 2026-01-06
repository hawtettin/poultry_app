from __future__ import annotations

import csv
from io import BytesIO

import openpyxl
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Q, Sum, Count
from django.db.models.functions import Coalesce
from django.forms import inlineformset_factory
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.cache import never_cache

from apps.core.models import Flock, House, Season
from apps.health.models import MortalityEvent, Treatment
from apps.auditlog.utils import log_event, snapshot
from apps.auditlog.models import AuditEvent, AccessLog
from apps.finance.models import Document, DocumentLine, Payment, Partner, Category, ExpenseAttachment

from .forms import (
    CreateSeriesForm,
    FlockEditForm,
    MortalityQuickAddForm,
    MortalityEditForm,
    SaleQuickAddForm,
    ExpenseDocumentForm,
    ExpenseLineForm,
    StaffUserCreateForm,
    PaymentEditForm,
)


WEEKDAYS_RO = ["luni", "marți", "miercuri", "joi", "vineri", "sâmbătă", "duminică"]

def is_manager(user) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name__in=["ADMIN", "MANAGER"]).exists()


def is_admin(user) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    return user.groups.filter(name="ADMIN").exists()


def can_modify_mortality(user, m: MortalityEvent) -> bool:
    if not user.is_authenticated:
        return False
    if user.is_superuser or is_manager(user):
        return True
    # employee: doar ce a creat el
    return (m.created_by_id is not None) and (m.created_by_id == user.id)


def can_modify_payment(user, p: Payment) -> bool:
    """Permisiuni UI pentru edit/ștergere/mark-paid la Payment.

    - ADMIN/MANAGER/superuser: orice plată.
    - EMPLOYEE: doar plățile create de el și doar în ziua curentă (după created_at).
    """

    if not getattr(user, "is_authenticated", False):
        return False

    if is_manager(user):
        return True

    # EMPLOYEE: doar ale lui
    if not user.groups.filter(name="EMPLOYEE").exists():
        return False

    if getattr(p, "created_by_id", None) != user.id:
        return False

    created_at = getattr(p, "created_at", None)
    if not created_at:
        return False

    try:
        created_day = timezone.localtime(created_at).date()
    except Exception:
        # dacă USE_TZ=False (datetime naive)
        created_day = created_at.date()
    return created_day == timezone.localdate()


def can_modify_sale(user, d: Document) -> bool:
    """Permisiuni UI pentru ștergere vânzare (Document doc_type='sale').

    - ADMIN/MANAGER/superuser: orice vânzare.
    - EMPLOYEE: doar vânzările create de el și doar în ziua curentă (după created_at).

    Notă: nu permitem ștergere pentru documente non-sale prin acest helper.
    """

    if not getattr(user, "is_authenticated", False):
        return False

    # doar pentru vânzări
    if getattr(d, "doc_type", None) != "sale":
        return False

    if is_manager(user):
        return True

    # EMPLOYEE: doar ale lui
    if not user.groups.filter(name="EMPLOYEE").exists():
        return False

    if getattr(d, "created_by_id", None) != user.id:
        return False

    created_at = getattr(d, "created_at", None)
    if not created_at:
        return False

    try:
        created_day = timezone.localtime(created_at).date()
    except Exception:
        created_day = created_at.date()

    return created_day == timezone.localdate()


def can_modify_expense(user, d: Document) -> bool:
    """Permisiuni UI pentru cheltuieli (Document doc_type='expense').

    - ADMIN/MANAGER/superuser: orice cheltuială.
    - EMPLOYEE: doar cheltuielile create de el și doar în ziua curentă.
    """

    if not getattr(user, "is_authenticated", False):
        return False

    if getattr(d, "doc_type", None) != "expense":
        return False

    if is_manager(user):
        return True

    # EMPLOYEE: doar ale lui
    if not user.groups.filter(name="EMPLOYEE").exists():
        return False

    if getattr(d, "created_by_id", None) != user.id:
        return False

    created_at = getattr(d, "created_at", None)
    if not created_at:
        return False

    try:
        created_day = timezone.localtime(created_at).date()
    except Exception:
        created_day = created_at.date()

    return created_day == timezone.localdate()


def safe_next_url(request, default: str) -> str:
    """Returnează un next URL safe (doar intern), altfel default."""

    nxt = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if nxt and url_has_allowed_host_and_scheme(
        url=nxt,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return nxt
    return default


@never_cache
@login_required
def dashboard(request):
    active_tab = (request.GET.get("tab") or "flocks").strip() or "flocks"

    # Folosim prefix-uri ca să evităm coliziuni (ex: ambele formulare au câmpul "date")
    mortality_form = MortalityQuickAddForm(prefix="mort")
    sale_form = SaleQuickAddForm(prefix="sale")

    if request.method == "POST":
        action = (request.POST.get("_action") or "").strip()

        # --------------------
        # Quick-add mortalitate
        # --------------------
        if action == "add_mortality":
            mortality_form = MortalityQuickAddForm(request.POST, prefix="mort")
            if mortality_form.is_valid():
                m = MortalityEvent.objects.create(
                    flock=mortality_form.cleaned_data["flock"],
                    date=mortality_form.cleaned_data["date"],
                    poultry_type=mortality_form.cleaned_data.get("poultry_type") or "white",
                    count=mortality_form.cleaned_data["count"],
                    reason=mortality_form.cleaned_data.get("reason", ""),
                    created_by=request.user,
                )
                log_event(
                    actor=request.user,
                    action="CREATE",
                    instance=m,
                    message=f"CREATE mortalitate: -{m.count} ({m.get_poultry_type_display()}) (lot {m.flock_id}) la {m.date}",
                    before=None,
                    after=snapshot(m),
                    request=request,
                )
                messages.success(request, "Mortalitatea a fost salvată (și s-a actualizat numărul curent).")
                return redirect(f"{request.path}?tab=mort")
            messages.error(request, "Nu am putut salva mortalitatea. Verifică datele din formular.")
            active_tab = "mort"

        # -------------
        # Quick-add sale
        # -------------
        elif action == "add_sale":
            sale_form = SaleQuickAddForm(request.POST, prefix="sale")
            if sale_form.is_valid():
                doc = sale_form.save(user=request.user)
                log_event(
                    actor=request.user,
                    action="CREATE",
                    instance=doc,
                    message=f"CREATE vânzare: doc#{doc.id} ({doc.total} {doc.currency})",
                    before=None,
                    after=snapshot(doc),
                    request=request,
                )
                messages.success(request, "Vânzarea a fost salvată.")
                return redirect(f"{request.path}?tab=sales")
            messages.error(request, "Nu am putut salva vânzarea. Verifică datele din formular.")
            active_tab = "sales"

    flocks = (
        Flock.objects.select_related("season", "house")
        .annotate(mortality_total=Coalesce(Sum("mortality_events__count"), 0))
        .order_by("-start_date", "-id")
    )

    # Mortalitate pe tip (pui albi / pui colorați)
    mort_rows = (
        MortalityEvent.objects
        .values("flock_id", "poultry_type")
        .annotate(s=Sum("count"))
    )
    mort_map = {
        (int(r["flock_id"]), str(r["poultry_type"])): int(r["s"] or 0)
        for r in mort_rows
    }

    # Vânzări pe tip (derivate din description; păstrăm compatibilitate cu datele existente)
    sold_white_rows = (
        DocumentLine.objects
        .filter(document__doc_type="sale", document__flock__isnull=False)
        .filter(description__iexact="Pui albi")
        .values("document__flock")
        .annotate(s=Sum("qty"))
    )
    sold_colored_rows = (
        DocumentLine.objects
        .filter(document__doc_type="sale", document__flock__isnull=False)
        .filter(Q(description__iexact="Pui colorați") | Q(description__iexact="Pui colorati"))
        .values("document__flock")
        .annotate(s=Sum("qty"))
    )

    sold_white_map = {int(r["document__flock"]): int(r["s"] or 0) for r in sold_white_rows}
    sold_colored_map = {int(r["document__flock"]): int(r["s"] or 0) for r in sold_colored_rows}

    for f in flocks:
        # inventar inițial pe tip (fallback pentru loturi vechi)
        init_white = int(getattr(f, "initial_white_count", 0) or 0)
        init_colored = int(getattr(f, "initial_colored_count", 0) or 0)
        if (init_white + init_colored) <= 0:
            init_white = int(getattr(f, "initial_count", 0) or 0)
            init_colored = 0
        init_total = int(init_white + init_colored)

        # mortalitate / vânzări pe tip
        mort_white = mort_map.get((int(f.id), "white"), 0)
        mort_colored = mort_map.get((int(f.id), "colored"), 0)
        sold_white = sold_white_map.get(int(f.id), 0)
        sold_colored = sold_colored_map.get(int(f.id), 0)

        # expunem în template
        f.initial_total = init_total
        f.initial_white = init_white
        f.initial_colored = init_colored

        f.mortality_white = mort_white
        f.mortality_colored = mort_colored
        f.mortality_total = int(mort_white + mort_colored)

        f.sold_white = sold_white
        f.sold_colored = sold_colored
        f.sold_total = int(sold_white + sold_colored)

        f.current_white = max(int(init_white) - int(mort_white) - int(sold_white), 0)
        f.current_colored = max(int(init_colored) - int(mort_colored) - int(sold_colored), 0)
        f.current_count = int(f.current_white + f.current_colored)

        f.mortality_pct = (100.0 * float(f.mortality_total) / float(init_total)) if init_total else 0.0

    recent = (
        MortalityEvent.objects.select_related("flock", "flock__season", "flock__house", "created_by")
        .order_by("-date", "-id")[:40]
    )
    # pregatim in template daca user poate modifica
    for m in recent:
        m.can_modify = can_modify_mortality(request.user, m)

    # -----------------
    # Vânzări (listare)
    # -----------------
    sales_from = parse_date((request.GET.get("sales_from") or "").strip())
    sales_to = parse_date((request.GET.get("sales_to") or "").strip())
    sales_buyer = (request.GET.get("sales_buyer") or "").strip()
    sales_flock = (request.GET.get("sales_flock") or "").strip()
    sales_only_debts = (request.GET.get("sales_only_debts") or "").strip() in ("1", "true", "True", "yes")
    sales_include_orphans = (request.GET.get("sales_include_orphans") or "").strip() in ("1", "true", "True", "yes")

    # Implicit, ascundem "orfanele" (vânzări care nu mai au plăți), dar le putem afișa la cerere.
    sales_qs = (
        Document.objects.filter(doc_type="sale")
        .select_related("flock", "flock__season", "flock__house", "partner")
        .prefetch_related("lines", "payments")
    )
    if not sales_include_orphans:
        sales_qs = sales_qs.filter(payments__isnull=False).distinct()
    if sales_from:
        sales_qs = sales_qs.filter(date__gte=sales_from)
    if sales_to:
        sales_qs = sales_qs.filter(date__lte=sales_to)
    if sales_buyer:
        sales_qs = sales_qs.filter(partner__name__icontains=sales_buyer)
    if sales_flock.isdigit():
        sales_qs = sales_qs.filter(flock_id=int(sales_flock))
    if sales_only_debts:
        sales_qs = sales_qs.filter(payments__status="due").distinct()

    sales_docs = list(sales_qs.order_by("-date", "-id")[:100])

    def _sum_qty(doc: Document, *, keys: set[str]) -> Decimal:
        total = Decimal("0")
        for ln in getattr(doc, "lines", []).all():
            desc = (ln.description or "").strip().lower()
            if desc in keys:
                total += (ln.qty or Decimal("0"))
        return total

    for d in sales_docs:
        d.qty_pui_albi = int(_sum_qty(d, keys={"pui albi"}) or 0)
        d.qty_pui_colorati = int(_sum_qty(d, keys={"pui colorați", "pui colorati"}) or 0)
        d.qty_furaj = _sum_qty(d, keys={"furaj"}).quantize(Decimal("0.001"))
        d.datorie = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
        d.bani = sum((p.amount for p in d.payments.all() if p.status == "paid"), Decimal("0.00")).quantize(Decimal("0.01"))
        try:
            d.is_orphan = len(list(d.payments.all())) == 0
        except Exception:
            d.is_orphan = False
        d.can_delete = can_modify_sale(request.user, d)

    # -----------------
    # Datorii pe cumpărător
    # -----------------
    debts_by_buyer = (
        Payment.objects.filter(status="due", document__doc_type="sale")
        .values("document__partner_id", "document__partner__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")[:30]
    )

    due_payments_qs = (
        Payment.objects.filter(status="due", document__doc_type="sale")
        .select_related(
            "document",
            "document__partner",
            "document__flock",
            "document__flock__season",
            "document__flock__house",
            "created_by",
        )
        .order_by("due_date", "id")[:80]
    )
    due_payments = list(due_payments_qs)
    for p in due_payments:
        p.can_modify = can_modify_payment(request.user, p)

    # -----------------
    # Raport rapid
    # -----------------
    today = timezone.localdate()
    report_from = sales_from or (today - timezone.timedelta(days=6) if hasattr(timezone, "timedelta") else today)
    report_to = sales_to or today
    # dacă nu există timezone.timedelta (în unele versiuni), importăm din datetime
    if not hasattr(timezone, "timedelta"):
        from datetime import timedelta
        report_from = sales_from or (today - timedelta(days=6))

    report_payments = Payment.objects.filter(
        document__doc_type="sale",
        document__date__gte=report_from,
        document__date__lte=report_to,
    )

    report_total_sales = (report_payments.aggregate(s=Sum("amount"))["s"] or Decimal("0.00")).quantize(Decimal("0.01"))
    report_total_debts = (
        report_payments.filter(status="due").aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    ).quantize(Decimal("0.01"))

    report_doc_ids = list(report_payments.values_list("document_id", flat=True).distinct())

    report_pui_albi = DocumentLine.objects.filter(document_id__in=report_doc_ids).filter(
        description__iexact="Pui albi"
    ).aggregate(s=Sum("qty"))["s"] or Decimal("0")

    report_pui_colorati = DocumentLine.objects.filter(document_id__in=report_doc_ids).filter(
        Q(description__iexact="Pui colorați") | Q(description__iexact="Pui colorati")
    ).aggregate(s=Sum("qty"))["s"] or Decimal("0")

    report_furaj = DocumentLine.objects.filter(document_id__in=report_doc_ids).filter(
        description__iexact="Furaj"
    ).aggregate(s=Sum("qty"))["s"] or Decimal("0")

    report_pui_total = int(report_pui_albi or 0) + int(report_pui_colorati or 0)
    report_furaj = (report_furaj or Decimal("0")).quantize(Decimal("0.001"))

    top_buyers = (
        report_payments.filter(status="paid")
        .values("document__partner__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")[:7]
    )

    top_debtors = (
        report_payments.filter(status="due")
        .values("document__partner__name")
        .annotate(total=Sum("amount"))
        .order_by("-total")[:7]
    )

    # -----------------
    # Cheltuieli (mini-ERP) - listare + rezumat
    # -----------------
    exp_from = parse_date((request.GET.get("exp_from") or "").strip())
    exp_to = parse_date((request.GET.get("exp_to") or "").strip())
    exp_house = (request.GET.get("exp_house") or "").strip()
    exp_flock = (request.GET.get("exp_flock") or "").strip()
    exp_supplier = (request.GET.get("exp_supplier") or "").strip()
    exp_status = (request.GET.get("exp_status") or "all").strip() or "all"
    exp_has_attach = (request.GET.get("exp_has_attach") or "").strip()
    exp_search = (request.GET.get("exp_search") or "").strip()

    expenses_qs = (
        Document.objects.filter(doc_type="expense")
        .select_related("season", "partner")
        .prefetch_related("lines", "payments", "attachments")
    )

    if exp_from:
        expenses_qs = expenses_qs.filter(date__gte=exp_from)
    if exp_to:
        expenses_qs = expenses_qs.filter(date__lte=exp_to)
    if exp_supplier:
        expenses_qs = expenses_qs.filter(partner__name__icontains=exp_supplier)

    # filtrare după alocări (linii)
    if exp_house.isdigit():
        expenses_qs = expenses_qs.filter(lines__house_id=int(exp_house)).distinct()
    if exp_flock.isdigit():
        expenses_qs = expenses_qs.filter(lines__flock_id=int(exp_flock)).distinct()

    if exp_search:
        expenses_qs = expenses_qs.filter(
            Q(doc_no__icontains=exp_search)
            | Q(notes__icontains=exp_search)
            | Q(lines__description__icontains=exp_search)
        ).distinct()

    if exp_has_attach in ("1", "true", "True", "yes"):
        expenses_qs = expenses_qs.filter(attachments__isnull=False).distinct()
    elif exp_has_attach in ("0", "false", "False", "no"):
        expenses_qs = expenses_qs.filter(attachments__isnull=True).distinct()

    expenses_qs = expenses_qs.annotate(
        paid_sum=Coalesce(Sum("payments__amount", filter=Q(payments__status="paid")), Decimal("0.00")),
        due_sum=Coalesce(Sum("payments__amount", filter=Q(payments__status="due")), Decimal("0.00")),
        attach_count=Count("attachments", distinct=True),
    )

    if exp_status == "paid":
        expenses_qs = expenses_qs.filter(due_sum=Decimal("0.00"), paid_sum__gt=Decimal("0.00"))
    elif exp_status == "unpaid":
        expenses_qs = expenses_qs.filter(paid_sum=Decimal("0.00"))
    elif exp_status == "partial":
        expenses_qs = expenses_qs.filter(paid_sum__gt=Decimal("0.00"), due_sum__gt=Decimal("0.00"))

    # ids (pentru totaluri fără dublări din join-uri)
    expense_doc_ids = list(expenses_qs.values_list("id", flat=True).distinct())
    expense_totals = {
        "total": Decimal("0.00"),
        "paid": Decimal("0.00"),
        "due": Decimal("0.00"),
        "count": len(expense_doc_ids),
    }
    if expense_doc_ids:
        expense_totals["total"] = (
            Document.objects.filter(id__in=expense_doc_ids).aggregate(s=Coalesce(Sum("total"), Decimal("0.00")))["s"]
            or Decimal("0.00")
        ).quantize(Decimal("0.01"))
        pay_aggr = Payment.objects.filter(document_id__in=expense_doc_ids).aggregate(
            paid=Coalesce(Sum("amount", filter=Q(status="paid")), Decimal("0.00")),
            due=Coalesce(Sum("amount", filter=Q(status="due")), Decimal("0.00")),
        )
        expense_totals["paid"] = (pay_aggr.get("paid") or Decimal("0.00")).quantize(Decimal("0.01"))
        expense_totals["due"] = (pay_aggr.get("due") or Decimal("0.00")).quantize(Decimal("0.01"))

    # total luna curentă
    month_start = today.replace(day=1)
    month_total_expenses = (
        Document.objects.filter(doc_type="expense", date__gte=month_start, date__lte=today)
        .aggregate(s=Coalesce(Sum("total"), Decimal("0.00")))["s"]
        or Decimal("0.00")
    ).quantize(Decimal("0.01"))

    # top denumiri (după liniile filtrate)
    top_expense_names = []
    if expense_doc_ids:
        top_expense_names = (
            DocumentLine.objects.filter(document_id__in=expense_doc_ids, document__doc_type="expense")
            .values("description")
            .annotate(total=Sum("line_total"))
            .order_by("-total")[:7]
        )

    expenses_docs = list(expenses_qs.order_by("-date", "-id")[:200])
    for d in expenses_docs:
        # status pe document (derivat din plăți)
        paid = (getattr(d, "paid_sum", None) or Decimal("0.00")).quantize(Decimal("0.01"))
        due = (getattr(d, "due_sum", None) or Decimal("0.00")).quantize(Decimal("0.01"))
        if paid > 0 and due > 0:
            d.pay_status = "partial"
            d.pay_status_label = "Parțial"
        elif due > 0 and paid == 0:
            d.pay_status = "unpaid"
            d.pay_status_label = "Neplătit"
        elif due == 0 and paid > 0:
            d.pay_status = "paid"
            d.pay_status_label = "Plătit"
        else:
            d.pay_status = "unpaid"
            d.pay_status_label = "Neplătit"

        # scadența următoare (din plăți due)
        try:
            dues = [p.due_date for p in d.payments.all() if p.status == "due" and p.due_date]
            d.next_due_date = min(dues) if dues else None
        except Exception:
            d.next_due_date = None

        # hale/loturi implicate (din linii)
        try:
            houses = []
            flocks_in_doc = []
            for ln in d.lines.all():
                if getattr(ln, "house", None):
                    houses.append(ln.house.name)
                if getattr(ln, "flock", None):
                    flocks_in_doc.append(ln.flock)
            d.houses_label = ", ".join(sorted(set(houses))) if houses else "—"
            d.flocks_count = len({f.id for f in flocks_in_doc}) if flocks_in_doc else 0
        except Exception:
            d.houses_label = "—"
            d.flocks_count = 0

        d.can_modify = can_modify_expense(request.user, d)

    # cost pe lot (defalcat) pentru filtrul curent
    expense_cost_by_flock = []
    if expense_doc_ids:
        rows = (
            DocumentLine.objects.filter(document_id__in=expense_doc_ids, document__doc_type="expense")
            .filter(flock__isnull=False)
            .values("flock")
            .annotate(cost=Sum("line_total"))
            .order_by("-cost")
        )
        flock_ids = [int(r["flock"]) for r in rows]
        flocks_map = {
            f.id: f
            for f in Flock.objects.filter(id__in=flock_ids).select_related("season", "house")
        }
        for r in rows:
            fid = int(r["flock"]) if r.get("flock") else None
            f = flocks_map.get(fid)
            cost = (r.get("cost") or Decimal("0.00")).quantize(Decimal("0.01"))
            if not f:
                continue
            init_total = int(getattr(f, "initial_count", 0) or 0)
            # folosim initial_total dacă există defalcare
            try:
                init_total = int((getattr(f, "initial_white_count", 0) or 0) + (getattr(f, "initial_colored_count", 0) or 0))
                if init_total <= 0:
                    init_total = int(getattr(f, "initial_count", 0) or 0)
            except Exception:
                init_total = int(getattr(f, "initial_count", 0) or 0)

            cost_per_chick = (cost / Decimal(init_total)).quantize(Decimal("0.0001")) if init_total else Decimal("0")
            expense_cost_by_flock.append({
                "flock": f,
                "cost": cost,
                "init_total": init_total,
                "cost_per_chick": cost_per_chick,
            })

    today_weekday = WEEKDAYS_RO[today.weekday()]

    return render(request, "ui/dashboard.html", {
        "flocks": flocks,
        "mortality_form": mortality_form,
        "sale_form": sale_form,
        "recent_mortalities": recent,
        "sales": sales_docs,
        "debts_by_buyer": debts_by_buyer,
        "due_payments": due_payments,
        "sales_from": sales_from,
        "sales_to": sales_to,
        "sales_buyer": sales_buyer,
        "sales_flock": sales_flock,
        "sales_only_debts": sales_only_debts,
        "sales_include_orphans": sales_include_orphans,
        "report_from": report_from,
        "report_to": report_to,
        "report_total_sales": report_total_sales,
        "report_total_debts": report_total_debts,
        "report_pui_total": report_pui_total,
        "report_furaj": report_furaj,
        "top_buyers": top_buyers,
        "top_debtors": top_debtors,
        "active_tab": active_tab,
        "today": today,
        "today_weekday": today_weekday,
        "is_manager": is_manager(request.user),

        # Cheltuieli
        "expenses": expenses_docs,
        "exp_from": exp_from,
        "exp_to": exp_to,
        "exp_house": exp_house,
        "exp_flock": exp_flock,
        "exp_supplier": exp_supplier,
        "exp_status": exp_status,
        "exp_has_attach": exp_has_attach,
        "exp_search": exp_search,
        "expense_totals": expense_totals,
        "month_total_expenses": month_total_expenses,
        "top_expense_names": top_expense_names,
        "expense_cost_by_flock": expense_cost_by_flock,
        "houses_list": House.objects.all().order_by("name"),
        "flocks_list": Flock.objects.select_related("season", "house").all().order_by("-start_date", "-id"),
    })


@login_required
def create_series(request):
    if not is_manager(request.user):
        messages.error(request, "Nu ai permisiuni să creezi serii/loturi. Cere acces de la administrator.")
        return redirect("ui:dashboard")

    if request.method == "POST":
        form = CreateSeriesForm(request.POST)
        if form.is_valid():
            season, flock = form.save()

            # audit (2 evenimente, pentru sezon si pentru flock)
            log_event(
                actor=request.user,
                action="CREATE",
                instance=season,
                message=f"CREATE sezon: {season.name}",
                after=snapshot(season),
                request=request,
            )
            log_event(
                actor=request.user,
                action="CREATE",
                instance=flock,
                message=f"CREATE lot: sezon={season.name}, hala={flock.house.name}, initial={flock.initial_count}",
                after=snapshot(flock),
                request=request,
            )

            messages.success(request, f"Serie creată: {season.name}. Lot nou în {flock.house.name} cu {flock.initial_count} capete.")
            return redirect("ui:dashboard")
        messages.error(request, "Nu am putut salva seria. Verifică formularul.")
    else:
        form = CreateSeriesForm()

    houses = House.objects.annotate(flock_count=Count("flocks")).order_by("name")

    return render(request, "ui/create_series.html", {"form": form, "houses": houses})


@login_required
def flock_edit(request, pk: int):
    """Editare lot: defalcare inițială pui albi / pui colorați.

    Doar MANAGER/ADMIN/superuser.
    """

    if not is_manager(request.user):
        messages.error(request, "Nu ai permisiuni să editezi loturi. Cere acces de la administrator.")
        return redirect("ui:dashboard")

    flock = get_object_or_404(Flock.objects.select_related("season", "house"), pk=pk)
    before = snapshot(flock)

    if request.method == "POST":
        form = FlockEditForm(request.POST, instance=flock)
        if form.is_valid():
            f2 = form.save()
            after = snapshot(f2)
            log_event(
                actor=request.user,
                action="UPDATE",
                instance=f2,
                message=(
                    f"UPDATE lot #{f2.id}: inițial_albi={before.get('initial_white_count')} -> {after.get('initial_white_count')}, "
                    f"inițial_colorați={before.get('initial_colored_count')} -> {after.get('initial_colored_count')}"
                ),
                before=before,
                after=after,
                request=request,
            )
            messages.success(request, "Lotul a fost actualizat.")
            return redirect("ui:dashboard")
        messages.error(request, "Nu am putut salva modificările. Verifică formularul.")
    else:
        form = FlockEditForm(instance=flock)

    return render(request, "ui/flock_edit.html", {"form": form, "flock": flock})



@login_required
def flock_delete(request, pk: int):
    """Șterge un lot (Flock) + toate datele asociate lui.

    Include:
      - Documente pe lot (vânzări / cheltuieli) + linii + plăți (cascade)
      - Mortalități + tratamente (cascade)
    """

    if not is_manager(request.user):
        messages.error(request, "Nu ai permisiuni să ștergi loturi. Cere acces de la administrator.")
        return redirect("ui:dashboard")

    flock = get_object_or_404(Flock.objects.select_related("season", "house"), pk=pk)

    default_next = reverse("ui:dashboard") + "?tab=flocks"
    next_url = safe_next_url(request, default_next)

    # Statistici pentru pagina de confirmare
    # Include și cheltuieli alocate prin linii (Document.flock poate fi NULL pentru cheltuieli "comune")
    docs_qs = Document.objects.filter(Q(flock=flock) | Q(lines__flock=flock)).distinct()
    docs_count = docs_qs.count()
    lines_count = DocumentLine.objects.filter(Q(document__flock=flock) | Q(flock=flock)).count()
    payments_count = Payment.objects.filter(document__in=docs_qs).count()
    mortality_count = MortalityEvent.objects.filter(flock=flock).count()
    treatments_count = Treatment.objects.filter(flock=flock).count()

    sold_white = (
        DocumentLine.objects
        .filter(document__doc_type="sale", document__flock=flock, description__iexact="Pui albi")
        .aggregate(s=Sum("qty"))["s"]
        or Decimal("0")
    )
    sold_colored = (
        DocumentLine.objects
        .filter(document__doc_type="sale", document__flock=flock)
        .filter(Q(description__iexact="Pui colorați") | Q(description__iexact="Pui colorati"))
        .aggregate(s=Sum("qty"))["s"]
        or Decimal("0")
    )
    sold_total = int(sold_white or 0) + int(sold_colored or 0)

    if request.method == "POST":
        before = snapshot(flock)
        season = flock.season
        house = flock.house

        log_event(
            actor=request.user,
            action="DELETE",
            instance=flock,
            message=f"DELETE lot #{flock.id}: {season.name} / {house.name}",
            before=before,
            after=None,
            request=request,
        )

        with transaction.atomic():
            # 1) ștergem documentele pe lot (Document.flock este PROTECT, deci trebuie înainte de flock.delete())
            docs_qs.delete()  # cascade: lines + payments
            # 2) ștergem lotul (cascade: mortalități + tratamente)
            flock.delete()

            # 3) dacă sezonul rămâne fără loturi și fără documente, îl ștergem (ca să poți recrea aceeași serie)
            try:
                if (not Flock.objects.filter(season=season).exists()) and (not Document.objects.filter(season=season).exists()):
                    season.delete()
            except Exception:
                pass

        messages.success(request, "Lotul a fost șters definitiv. (Stocurile / vânzările se recalculează automat)")
        return redirect(next_url)

    return render(
        request,
        "ui/flock_confirm_delete.html",
        {
            "flock": flock,
            "next_url": next_url,
            "docs_count": docs_count,
            "lines_count": lines_count,
            "payments_count": payments_count,
            "mortality_count": mortality_count,
            "treatments_count": treatments_count,
            "sold_white": int(sold_white or 0),
            "sold_colored": int(sold_colored or 0),
            "sold_total": sold_total,
        },
    )


@login_required
def house_delete(request, pk: int):
    """Șterge o hală (House) + toate loturile din ea + datele aferente.

    Atenție: acțiune ireversibilă.
    """

    if not is_manager(request.user):
        messages.error(request, "Nu ai permisiuni să ștergi hale. Cere acces de la administrator.")
        return redirect("ui:dashboard")

    house = get_object_or_404(House, pk=pk)

    default_next = reverse("ui:create_series")
    next_url = safe_next_url(request, default_next)

    flocks = list(Flock.objects.filter(house=house).select_related("season"))

    # Statistici pentru confirmare
    flocks_count = len(flocks)
    docs_qs = Document.objects.filter(Q(flock__house=house) | Q(lines__house=house)).distinct()
    docs_count = docs_qs.count()
    payments_count = Payment.objects.filter(document__in=docs_qs).count()
    lines_count = DocumentLine.objects.filter(Q(document__flock__house=house) | Q(house=house)).count()
    mortality_count = MortalityEvent.objects.filter(flock__house=house).count()
    treatments_count = Treatment.objects.filter(flock__house=house).count()

    if request.method == "POST":
        before = snapshot(house)

        log_event(
            actor=request.user,
            action="DELETE",
            instance=house,
            message=f"DELETE hală #{house.id}: {house.name} (și toate loturile din hală)",
            before=before,
            after=None,
            request=request,
        )

        with transaction.atomic():
            # 1) Ștergem toate documentele alocate halei (atât prin Document.flock, cât și prin DocumentLine.house)
            docs_qs.delete()  # cascade: lines + payments + attachments

            # 2) Ștergem loturile (cascade: mortalitate + tratamente)
            seasons = set()
            for f in flocks:
                seasons.add(f.season_id)
                f.delete()

            # 3) Ștergem sezoanele rămase goale
            for season_id in seasons:
                try:
                    if (not Flock.objects.filter(season_id=season_id).exists()) and (not Document.objects.filter(season_id=season_id).exists()):
                        Season.objects.filter(id=season_id).delete()
                except Exception:
                    pass

            # 4) În final, ștergem hala
            house.delete()

        messages.success(request, "Hala a fost ștearsă definitiv (împreună cu toate loturile și documentele din ea).")
        return redirect(next_url)

    return render(
        request,
        "ui/house_confirm_delete.html",
        {
            "house": house,
            "flocks": flocks,
            "next_url": next_url,
            "flocks_count": flocks_count,
            "docs_count": docs_count,
            "lines_count": lines_count,
            "payments_count": payments_count,
            "mortality_count": mortality_count,
            "treatments_count": treatments_count,
        },
    )


@login_required
def cleanup_all(request):
    """Curățare totală: șterge toate datele operaționale (loturi, hale, vânzări, mortalitate, etc).

    Nu șterge utilizatorii/grupurile (ca să nu te blochezi din aplicație).
    """

    if not is_manager(request.user):
        messages.error(request, "Nu ai permisiuni să faci curățare totală. Cere acces de la administrator.")
        return redirect("ui:dashboard")

    default_next = reverse("ui:dashboard")
    next_url = safe_next_url(request, default_next)

    # Statistici pentru confirmare
    counts = {
        "houses": House.objects.count(),
        "seasons": Season.objects.count(),
        "flocks": Flock.objects.count(),
        "mortalities": MortalityEvent.objects.count(),
        "treatments": Treatment.objects.count(),
        "documents": Document.objects.count(),
        "document_lines": DocumentLine.objects.count(),
        "payments": Payment.objects.count(),
        "expense_attachments": ExpenseAttachment.objects.count(),
        "partners": Partner.objects.count(),
        "categories": Category.objects.count(),
        "audit_events": AuditEvent.objects.count(),
        "access_logs": AccessLog.objects.count(),
    }

    if request.method == "POST":
        confirm_text = (request.POST.get("confirm_text") or "").strip().upper()
        if confirm_text not in ("STERGE TOT", "ȘTERGE TOT", "DELETE ALL", "DELETE"):
            messages.error(request, "Confirmare greșită. Scrie exact: STERGE TOT")
            return render(request, "ui/cleanup_confirm.html", {"counts": counts, "next_url": next_url})

        # Notă: nu logăm în audit aici, pentru că audit-ul se șterge (site clean).
        with transaction.atomic():
            # 1) Documente (cascade: plăți + linii)
            Document.objects.all().delete()

            # 2) Sănătate / loturi (mortalități și tratamente oricum sunt CASCADE la ștergere de lot)
            MortalityEvent.objects.all().delete()
            Treatment.objects.all().delete()
            Flock.objects.all().delete()

            # 3) Serii / hale
            Season.objects.all().delete()
            House.objects.all().delete()

            # 4) Parteneri / categorii
            Partner.objects.all().delete()
            Category.objects.all().delete()

            # 5) Audit / accesări
            AuditEvent.objects.all().delete()
            AccessLog.objects.all().delete()

        messages.success(request, "Curățare totală completă: toate datele au fost șterse. Aplicația este acum 'clean'.")
        return redirect(next_url)

    return render(request, "ui/cleanup_confirm.html", {"counts": counts, "next_url": next_url})


@login_required
def users_list(request):
    """Listă utilizatori (angajați + manageri).

    Doar ADMIN (sau superuser) poate vedea această pagină.
    """

    if not is_admin(request.user):
        messages.error(request, "Nu ai permisiuni să vezi lista de utilizatori.")
        return redirect("ui:dashboard")

    User = get_user_model()

    qs = (
        User.objects.filter(groups__name__in=["EMPLOYEE", "MANAGER"])
        .distinct()
        .order_by("username")
    )
    users = list(qs)

    # adăugăm un label simplu pentru rol (pentru tabel)
    for u in users:
        if getattr(u, "is_superuser", False):
            u.role_label = "Admin"
        elif u.groups.filter(name="MANAGER").exists():
            u.role_label = "Manager fermă"
        elif u.groups.filter(name="EMPLOYEE").exists():
            u.role_label = "Angajat"
        else:
            u.role_label = "-"

    return render(request, "ui/users_list.html", {"employees": users, "is_manager": is_manager(request.user)})


@login_required
def user_create(request):
    """Creează cont pentru EMPLOYEE sau MANAGER.

    Doar ADMIN (sau superuser) poate crea conturi.
    """

    if not is_admin(request.user):
        messages.error(request, "Doar ADMIN poate crea conturi.")
        return redirect("ui:dashboard")

    if request.method == "POST":
        form = StaffUserCreateForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    u = form.save(commit=True)
            except IntegrityError:
                form.add_error("username", "Acest username există deja. Alege altul (ex: vasile1).")
                messages.error(request, "Username duplicat.")
                return render(request, "ui/user_create.html", {"form": form})

            log_event(
                actor=request.user,
                action="CREATE",
                instance=u,
                message=f"CREATE utilizator: {u.username}",
                after=snapshot(u),
                request=request,
            )
            messages.success(request, f"Utilizator creat: {u.username}")
            return redirect("ui:users_list")

        messages.error(request, "Nu am putut crea utilizatorul. Verifică formularul.")
    else:
        form = StaffUserCreateForm()

    return render(request, "ui/user_create.html", {"form": form})


@login_required
def mortality_edit(request, pk: int):
    m = get_object_or_404(MortalityEvent.objects.select_related("flock", "created_by"), pk=pk)
    if not can_modify_mortality(request.user, m):
        messages.error(request, "Nu ai drepturi să editezi această înregistrare.")
        return redirect("ui:dashboard")

    before = snapshot(m)

    if request.method == "POST":
        form = MortalityEditForm(request.POST, instance=m)
        if form.is_valid():
            m2 = form.save()
            after = snapshot(m2)
            log_event(
                actor=request.user,
                action="UPDATE",
                instance=m2,
                message=f"UPDATE mortalitate #{m2.id}: {before.get('count')} -> {after.get('count')}",
                before=before,
                after=after,
                request=request,
            )
            messages.success(request, "Mortalitatea a fost actualizată.")
            return redirect("ui:dashboard")
        messages.error(request, "Nu am putut salva modificările. Verifică formularul.")
    else:
        form = MortalityEditForm(instance=m)

    return render(request, "ui/mortality_edit.html", {"form": form, "m": m})


@login_required
def mortality_delete(request, pk: int):
    m = get_object_or_404(MortalityEvent.objects.select_related("flock", "created_by"), pk=pk)
    if not can_modify_mortality(request.user, m):
        messages.error(request, "Nu ai drepturi să ștergi această înregistrare.")
        return redirect("ui:dashboard")

    if request.method == "POST":
        before = snapshot(m)
        mid = m.id
        m.delete()
        # log dupa delete (pastram object_id si inainte)
        class Dummy:  # doar ca sa avem meta label ok
            _meta = type("M", (), {"app_label": "health", "model_name": "mortalityevent"})()
            pk = mid
        dummy = Dummy()
        log_event(
            actor=request.user,
            action="DELETE",
            instance=dummy,
            message=f"DELETE mortalitate #{mid}: -{before.get('count')} (lot {before.get('flock')}) la {before.get('date')}",
            before=before,
            after=None,
            request=request,
        )
        messages.success(request, "Înregistrarea a fost ștearsă.")
        return redirect("ui:dashboard")

    return render(request, "ui/mortality_confirm_delete.html", {"m": m})


@login_required
def history(request):
    """Istoric schimbări (audit CRUD).

    - EMPLOYEE: vede doar istoricul lui
    - MANAGER/ADMIN: vede tot + poate filtra după utilizator și zi
    """

    mgr = is_manager(request.user)
    User = get_user_model()

    selected_user = (request.GET.get("user") or "").strip()
    day = parse_date((request.GET.get("day") or "").strip())

    qs = AuditEvent.objects.select_related("actor").all()

    if not mgr:
        qs = qs.filter(actor=request.user)
    else:
        if selected_user.isdigit():
            qs = qs.filter(actor_id=int(selected_user))
        if day:
            qs = qs.filter(created_at__date=day)

    qs = qs.order_by("-created_at", "-id")[:300]

    employees = []
    if mgr:
        employees = (
            User.objects.filter(groups__name__in=["EMPLOYEE", "MANAGER"])
            .distinct()
            .order_by("username")
        )

    return render(
        request,
        "ui/history.html",
        {
            "events": qs,
            "is_manager": mgr,
            "employees": employees,
            "selected_user": selected_user,
            "selected_day": (day.isoformat() if day else ""),
        },
    )


@login_required
def access_history(request):
    """Istoric accesări (LOGIN/LOGOUT)."""

    from apps.auditlog.models import AccessLog

    mgr = is_manager(request.user)
    User = get_user_model()

    selected_user = (request.GET.get("user") or "").strip()
    day = parse_date((request.GET.get("day") or "").strip())

    qs = AccessLog.objects.select_related("actor").all()
    if not mgr:
        qs = qs.filter(actor=request.user)
    else:
        if selected_user.isdigit():
            qs = qs.filter(actor_id=int(selected_user))
        if day:
            qs = qs.filter(created_at__date=day)

    qs = qs.order_by("-created_at", "-id")[:400]

    employees = []
    if mgr:
        employees = (
            User.objects.filter(groups__name__in=["EMPLOYEE", "MANAGER"])
            .distinct()
            .order_by("username")
        )

    return render(
        request,
        "ui/access.html",
        {
            "events": qs,
            "is_manager": mgr,
            "employees": employees,
            "selected_user": selected_user,
            "selected_day": (day.isoformat() if day else ""),
        },
    )


@never_cache
@login_required
def payment_ledger(request):
    """Ledger (registru) plăți.

    Vizibilitate:
    - ADMIN/MANAGER: toate plățile
    - EMPLOYEE: doar plățile create de el

    Acțiuni (edit/delete):
    - ADMIN/MANAGER: orice plată
    - EMPLOYEE: doar plățile lui, doar în ziua curentă (după created_at)

    Filtre (GET):
      - status: all|due|paid
      - due_from / due_to (YYYY-MM-DD)  (după due_date)
      - paid_from / paid_to (YYYY-MM-DD) (după paid_date)
      - buyer (string)  (document.partner.name)
      - user (id) (doar manager/admin)
    """

    mgr = is_manager(request.user)
    User = get_user_model()

    status = (request.GET.get("status") or "all").strip() or "all"
    due_from = parse_date((request.GET.get("due_from") or "").strip())
    due_to = parse_date((request.GET.get("due_to") or "").strip())
    paid_from = parse_date((request.GET.get("paid_from") or "").strip())
    paid_to = parse_date((request.GET.get("paid_to") or "").strip())
    buyer = (request.GET.get("buyer") or "").strip()
    selected_user = (request.GET.get("user") or "").strip()

    qs = (
        Payment.objects.select_related("document", "document__partner", "created_by")
        .all()
    )

    if not mgr:
        qs = qs.filter(created_by=request.user)
    else:
        if selected_user.isdigit():
            qs = qs.filter(created_by_id=int(selected_user))

    if status in ("due", "paid"):
        qs = qs.filter(status=status)

    if due_from:
        qs = qs.filter(due_date__gte=due_from)
    if due_to:
        qs = qs.filter(due_date__lte=due_to)

    if paid_from:
        qs = qs.filter(paid_date__gte=paid_from)
    if paid_to:
        qs = qs.filter(paid_date__lte=paid_to)

    if buyer:
        qs = qs.filter(document__partner__name__icontains=buyer)

    qs = qs.order_by("-due_date", "-id")

    # Totals pentru filtrul curent
    totals = qs.aggregate(
        total=Coalesce(Sum("amount"), Decimal("0.00")),
        total_due=Coalesce(Sum("amount", filter=Q(status="due")), Decimal("0.00")),
        total_paid=Coalesce(Sum("amount", filter=Q(status="paid")), Decimal("0.00")),
    )

    payments = list(qs[:600])
    for p in payments:
        p.can_modify = can_modify_payment(request.user, p)
        # label simplu pentru tabel
        if p.created_by_id:
            p.created_by_label = getattr(p.created_by, "username", "") or ""
        else:
            p.created_by_label = "-"

    employees = []
    if mgr:
        employees = (
            User.objects.filter(groups__name__in=["EMPLOYEE", "MANAGER"]).distinct().order_by("username")
        )

    return render(
        request,
        "ui/payment_ledger.html",
        {
            "payments": payments,
            "is_manager": mgr,
            "employees": employees,
            "status": status,
            "due_from": (due_from.isoformat() if due_from else ""),
            "due_to": (due_to.isoformat() if due_to else ""),
            "paid_from": (paid_from.isoformat() if paid_from else ""),
            "paid_to": (paid_to.isoformat() if paid_to else ""),
            "buyer": buyer,
            "selected_user": selected_user,
            "totals": totals,
            "today": timezone.localdate().isoformat(),
        },
    )


@login_required
def payment_edit(request, pk: int):
    p = get_object_or_404(
        Payment.objects.select_related("document", "document__partner", "created_by"),
        pk=pk,
    )

    if not can_modify_payment(request.user, p):
        messages.error(request, "Nu ai drepturi să editezi această plată.")
        return redirect("ui:payment_ledger")

    before = snapshot(p)

    next_url = (request.GET.get("next") or "").strip() or reverse("ui:payment_ledger")

    if request.method == "POST":
        form = PaymentEditForm(request.POST, instance=p)
        if form.is_valid():
            p2 = form.save()

            log_event(
                actor=request.user,
                action="UPDATE",
                instance=p2,
                message=f"UPDATE payment: #{p2.id} ({p2.amount} {p2.document.currency}) status={p2.status}",
                before=before,
                after=snapshot(p2),
                request=request,
            )
            messages.success(request, "Plata a fost actualizată.")
            return redirect(next_url)

        messages.error(request, "Nu am putut salva plata. Verifică datele.")
    else:
        form = PaymentEditForm(instance=p)

    return render(
        request,
        "ui/payment_edit.html",
        {
            "form": form,
            "p": p,
            "next_url": next_url,
        },
    )


@login_required
def payment_delete(request, pk: int):
    p = get_object_or_404(
        Payment.objects.select_related("document", "document__partner", "created_by"),
        pk=pk,
    )

    if not can_modify_payment(request.user, p):
        messages.error(request, "Nu ai drepturi să ștergi această plată.")
        return redirect("ui:payment_ledger")

    next_url = (request.GET.get("next") or "").strip() or reverse("ui:payment_ledger")

    # Pentru UI: ștergerea plății nu înseamnă neapărat ștergerea vânzării.
    # Dacă user-ul vrea să anuleze vânzarea (și să se refacă Nr vânduți / Nr curent),
    # trebuie șters Document-ul.
    can_delete_sale = False
    if getattr(p, "document", None) is not None:
        try:
            can_delete_sale = can_modify_sale(request.user, p.document)
        except Exception:
            can_delete_sale = False

    if request.method == "POST":
        before = snapshot(p)
        log_event(
            actor=request.user,
            action="DELETE",
            instance=p,
            message=f"DELETE payment: #{p.id} ({p.amount} {p.document.currency}) status={p.status}",
            before=before,
            after=None,
            request=request,
        )
        p.delete()
        messages.success(request, "Plata a fost ștearsă.")
        return redirect(next_url)

    return render(
        request,
        "ui/payment_confirm_delete.html",
        {
            "p": p,
            "next_url": next_url,
            "can_delete_sale": can_delete_sale,
        },
    )


@login_required
def sale_delete(request, pk: int):
    """Șterge o vânzare (Document doc_type='sale') + liniile + plățile asociate.

    Important: pentru a actualiza "Nr vânduți" / "Nr curent", trebuie ștearsă vânzarea
    (Document + DocumentLine). Ștergerea unei plăți (Payment) nu elimină automat cantitățile vândute.
    """

    d = get_object_or_404(
        Document.objects
        .select_related("flock", "flock__season", "flock__house", "partner", "created_by")
        .prefetch_related("lines", "payments"),
        pk=pk,
        doc_type="sale",
    )

    if not can_modify_sale(request.user, d):
        messages.error(request, "Nu ai drepturi să ștergi această vânzare.")
        return redirect("ui:dashboard")

    default_next = reverse("ui:dashboard") + "?tab=sales"
    next_url = safe_next_url(request, default_next)

    if request.method == "POST":
        before = snapshot(d)

        buyer = ""
        try:
            buyer = d.partner.name if d.partner else ""
        except Exception:
            buyer = ""

        log_event(
            actor=request.user,
            action="DELETE",
            instance=d,
            message=f"DELETE vânzare: doc#{d.id} ({d.total} {d.currency}) buyer={buyer}",
            before=before,
            after=None,
            request=request,
        )

        # cascade: DocumentLine + Payment
        d.delete()
        messages.success(request, "Vânzarea a fost ștearsă. (Nr vânduți / Nr curent se recalculează automat)")
        return redirect(next_url)

    def _sum_qty(doc: Document, *, keys: set[str]) -> Decimal:
        total = Decimal("0")
        for ln in doc.lines.all():
            desc = (ln.description or "").strip().lower()
            if desc in keys:
                total += (ln.qty or Decimal("0"))
        return total

    d.qty_pui_albi = int(_sum_qty(d, keys={"pui albi"}) or 0)
    d.qty_pui_colorati = int(_sum_qty(d, keys={"pui colorați", "pui colorati"}) or 0)
    d.qty_furaj = _sum_qty(d, keys={"furaj"}).quantize(Decimal("0.001"))
    d.datorie = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
    d.bani = sum((p.amount for p in d.payments.all() if p.status == "paid"), Decimal("0.00")).quantize(Decimal("0.01"))

    return render(
        request,
        "ui/sale_confirm_delete.html",
        {
            "d": d,
            "next_url": next_url,
        },
    )


# -----------------
# Cheltuieli (mini-ERP)
# -----------------


@login_required
def expense_create(request):
    """Creare cheltuială (Document doc_type='expense') + linii alocate + plăți + atașamente."""

    default_next = reverse("ui:dashboard") + "?tab=expenses"
    next_url = safe_next_url(request, default_next)

    LineFormSet = inlineformset_factory(
        Document,
        DocumentLine,
        form=ExpenseLineForm,
        extra=8,
        can_delete=True,
    )

    if request.method == "POST":
        doc_form = ExpenseDocumentForm(request.POST, request.FILES)

        # instanță dummy doar pentru validare formset (înainte de a avea Document salvat)
        dummy = Document(doc_type="expense", status="approved", currency="RON")
        formset = LineFormSet(request.POST, instance=dummy, prefix="lines")

        if doc_form.is_valid() and formset.is_valid():
            # măcar o linie completată
            has_any_line = False
            for f in formset.forms:
                try:
                    cd = f.cleaned_data
                except Exception:
                    cd = None
                if not cd:
                    continue
                if cd.get("DELETE"):
                    continue
                # dacă form-ul nu e gol, sigur e o linie
                if cd.get("description"):
                    has_any_line = True
                    break
            if not has_any_line:
                messages.error(request, "Adaugă cel puțin o linie de cheltuială.")
                return render(request, "ui/expense_form.html", {"doc_form": doc_form, "formset": formset, "next_url": next_url})

            with transaction.atomic():
                supplier_name = doc_form.cleaned_data["supplier_name"]
                partner, _ = Partner.objects.get_or_create(
                    name=supplier_name,
                    defaults={"partner_type": "supplier"},
                )

                season = doc_form.cleaned_data["season"]

                doc = Document.objects.create(
                    doc_type="expense",
                    status="approved",
                    season=season,
                    flock=None,
                    partner=partner,
                    doc_no=(doc_form.cleaned_data.get("doc_no") or ""),
                    date=doc_form.cleaned_data["date"],
                    currency="RON",
                    vat_rate=(doc_form.cleaned_data.get("vat_rate") or Decimal("0")),
                    notes=(doc_form.cleaned_data.get("notes") or ""),
                    created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                )

                # salvăm liniile
                formset.instance = doc
                formset.save()

                # Totale (TVA inclus)
                doc.recalc_totals()
                doc.save(update_fields=["subtotal", "vat", "total", "vat_rate"])

                # Plăți inițiale (conform status)
                doc.payments.all().delete()
                status = (doc_form.cleaned_data.get("payment_status") or "unpaid").strip() or "unpaid"
                method = (doc_form.cleaned_data.get("payment_method") or "bank").strip() or "bank"
                due_date = doc_form.cleaned_data.get("due_date") or doc.date
                paid_amount = (doc_form.cleaned_data.get("paid_amount") or Decimal("0.00")).quantize(Decimal("0.01"))
                total = (doc.total or Decimal("0.00")).quantize(Decimal("0.01"))

                # normalize paid_amount
                if paid_amount < 0:
                    paid_amount = Decimal("0.00")
                if paid_amount > total:
                    paid_amount = total

                if status == "paid":
                    if total > 0:
                        Payment.objects.create(
                            document=doc,
                            due_date=doc.date,
                            paid_date=doc.date,
                            amount=total,
                            method=method,
                            status="paid",
                            created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                        )
                elif status == "partial":
                    if paid_amount > 0:
                        Payment.objects.create(
                            document=doc,
                            due_date=doc.date,
                            paid_date=doc.date,
                            amount=paid_amount,
                            method=method,
                            status="paid",
                            created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                        )
                    remaining = (total - paid_amount).quantize(Decimal("0.01"))
                    if remaining > 0:
                        Payment.objects.create(
                            document=doc,
                            due_date=due_date,
                            amount=remaining,
                            method=method,
                            status="due",
                            created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                        )
                else:
                    # unpaid
                    if total > 0:
                        Payment.objects.create(
                            document=doc,
                            due_date=due_date,
                            amount=total,
                            method=method,
                            status="due",
                            created_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                        )

                # Atașamente multiple
                for f in request.FILES.getlist("attachments"):
                    try:
                        ExpenseAttachment.objects.create(
                            document=doc,
                            file=f,
                            original_name=getattr(f, "name", "") or "",
                            uploaded_by=request.user if getattr(request.user, "is_authenticated", False) else None,
                        )
                    except Exception:
                        pass

                log_event(
                    actor=request.user,
                    action="CREATE",
                    instance=doc,
                    message=f"CREATE cheltuială: doc#{doc.id} ({doc.total} {doc.currency})",
                    before=None,
                    after=snapshot(doc),
                    request=request,
                )

            messages.success(request, "Cheltuiala a fost salvată.")
            return redirect(next_url)

        messages.error(request, "Nu am putut salva cheltuiala. Verifică datele.")

    else:
        doc_form = ExpenseDocumentForm()
        formset = LineFormSet(prefix="lines", instance=Document())

    return render(request, "ui/expense_form.html", {"doc_form": doc_form, "formset": formset, "next_url": next_url})


@login_required
def expense_detail(request, pk: int):
    d = get_object_or_404(
        Document.objects
        .select_related("season", "partner", "created_by")
        .prefetch_related("lines", "payments", "attachments"),
        pk=pk,
        doc_type="expense",
    )

    d.can_modify = can_modify_expense(request.user, d)

    # status derivat
    paid = sum((p.amount for p in d.payments.all() if p.status == "paid"), Decimal("0.00")).quantize(Decimal("0.01"))
    due = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
    if paid > 0 and due > 0:
        d.pay_status_label = "Parțial"
    elif due > 0 and paid == 0:
        d.pay_status_label = "Neplătit"
    elif due == 0 and paid > 0:
        d.pay_status_label = "Plătit"
    else:
        d.pay_status_label = "Neplătit"

    return render(request, "ui/expense_detail.html", {"d": d})


@login_required
def expense_edit(request, pk: int):
    d = get_object_or_404(
        Document.objects
        .select_related("season", "partner", "created_by")
        .prefetch_related("lines", "payments", "attachments"),
        pk=pk,
        doc_type="expense",
    )

    if not can_modify_expense(request.user, d):
        messages.error(request, "Nu ai drepturi să editezi această cheltuială.")
        return redirect("ui:dashboard")

    default_next = reverse("ui:expense_detail", args=[d.id])
    next_url = safe_next_url(request, default_next)

    LineFormSet = inlineformset_factory(
        Document,
        DocumentLine,
        form=ExpenseLineForm,
        extra=5,
        can_delete=True,
    )

    # status inițial pentru UI
    paid_sum = sum((p.amount for p in d.payments.all() if p.status == "paid"), Decimal("0.00")).quantize(Decimal("0.01"))
    due_sum = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
    if paid_sum > 0 and due_sum > 0:
        init_status = "partial"
        init_paid_amount = paid_sum
        init_due_date = min([p.due_date for p in d.payments.all() if p.status == "due" and p.due_date] or [d.date])
    elif due_sum > 0 and paid_sum == 0:
        init_status = "unpaid"
        init_paid_amount = Decimal("0.00")
        init_due_date = min([p.due_date for p in d.payments.all() if p.status == "due" and p.due_date] or [d.date])
    else:
        init_status = "paid"
        init_paid_amount = (d.total or Decimal("0.00")).quantize(Decimal("0.01"))
        init_due_date = d.date

    if request.method == "POST":
        doc_form = ExpenseDocumentForm(request.POST, request.FILES)
        formset = LineFormSet(request.POST, instance=d, prefix="lines")

        if doc_form.is_valid() and formset.is_valid():
            before = snapshot(d)
            with transaction.atomic():
                supplier_name = doc_form.cleaned_data["supplier_name"]
                partner, _ = Partner.objects.get_or_create(
                    name=supplier_name,
                    defaults={"partner_type": "supplier"},
                )
                d.partner = partner
                d.season = doc_form.cleaned_data["season"]
                d.doc_no = (doc_form.cleaned_data.get("doc_no") or "")
                d.date = doc_form.cleaned_data["date"]
                d.vat_rate = (doc_form.cleaned_data.get("vat_rate") or Decimal("0"))
                d.notes = (doc_form.cleaned_data.get("notes") or "")
                d.save(update_fields=["partner", "season", "doc_no", "date", "vat_rate", "notes"])

                formset.save()

                d.recalc_totals()
                d.save(update_fields=["subtotal", "vat", "total"])

                # recreăm plățile
                d.payments.all().delete()
                status = (doc_form.cleaned_data.get("payment_status") or "unpaid").strip() or "unpaid"
                method = (doc_form.cleaned_data.get("payment_method") or "bank").strip() or "bank"
                due_date = doc_form.cleaned_data.get("due_date") or d.date
                paid_amount = (doc_form.cleaned_data.get("paid_amount") or Decimal("0.00")).quantize(Decimal("0.01"))
                total = (d.total or Decimal("0.00")).quantize(Decimal("0.01"))
                if paid_amount < 0:
                    paid_amount = Decimal("0.00")
                if paid_amount > total:
                    paid_amount = total

                if status == "paid":
                    if total > 0:
                        Payment.objects.create(
                            document=d,
                            due_date=d.date,
                            paid_date=d.date,
                            amount=total,
                            method=method,
                            status="paid",
                            created_by=request.user,
                        )
                elif status == "partial":
                    if paid_amount > 0:
                        Payment.objects.create(
                            document=d,
                            due_date=d.date,
                            paid_date=d.date,
                            amount=paid_amount,
                            method=method,
                            status="paid",
                            created_by=request.user,
                        )
                    remaining = (total - paid_amount).quantize(Decimal("0.01"))
                    if remaining > 0:
                        Payment.objects.create(
                            document=d,
                            due_date=due_date,
                            amount=remaining,
                            method=method,
                            status="due",
                            created_by=request.user,
                        )
                else:
                    if total > 0:
                        Payment.objects.create(
                            document=d,
                            due_date=due_date,
                            amount=total,
                            method=method,
                            status="due",
                            created_by=request.user,
                        )

                # atașamente noi (nu ștergem pe cele vechi automat)
                for f in request.FILES.getlist("attachments"):
                    try:
                        ExpenseAttachment.objects.create(
                            document=d,
                            file=f,
                            original_name=getattr(f, "name", "") or "",
                            uploaded_by=request.user,
                        )
                    except Exception:
                        pass

                log_event(
                    actor=request.user,
                    action="UPDATE",
                    instance=d,
                    message=f"UPDATE cheltuială: doc#{d.id}",
                    before=before,
                    after=snapshot(d),
                    request=request,
                )

            messages.success(request, "Cheltuiala a fost actualizată.")
            return redirect(next_url)

        messages.error(request, "Nu am putut salva modificările. Verifică datele.")

    else:
        doc_form = ExpenseDocumentForm(initial={
            "date": d.date,
            "season": d.season,
            "supplier_name": (d.partner.name if d.partner else ""),
            "doc_no": d.doc_no,
            "vat_rate": d.vat_rate,
            "notes": d.notes,
            "payment_status": init_status,
            "payment_method": (d.payments.first().method if d.payments.exists() else "bank"),
            "due_date": init_due_date,
            "paid_amount": init_paid_amount,
        })
        formset = LineFormSet(prefix="lines", instance=d)

    return render(request, "ui/expense_form.html", {"doc_form": doc_form, "formset": formset, "doc": d, "next_url": next_url})


@login_required
def expense_delete(request, pk: int):
    d = get_object_or_404(
        Document.objects.select_related("season", "partner", "created_by").prefetch_related("lines", "payments", "attachments"),
        pk=pk,
        doc_type="expense",
    )

    if not can_modify_expense(request.user, d):
        messages.error(request, "Nu ai drepturi să ștergi această cheltuială.")
        return redirect("ui:dashboard")

    default_next = reverse("ui:dashboard") + "?tab=expenses"
    next_url = safe_next_url(request, default_next)

    if request.method == "POST":
        before = snapshot(d)
        log_event(
            actor=request.user,
            action="DELETE",
            instance=d,
            message=f"DELETE cheltuială: doc#{d.id} ({d.total} {d.currency})",
            before=before,
            after=None,
            request=request,
        )
        d.delete()
        messages.success(request, "Cheltuiala a fost ștearsă definitiv.")
        return redirect(next_url)

    return render(
        request,
        "ui/expense_confirm_delete.html",
        {
            "d": d,
            "next_url": next_url,
            "lines_count": d.lines.count(),
            "payments_count": d.payments.count(),
            "attachments_count": d.attachments.count(),
        },
    )


@login_required
def expense_attachment_delete(request, pk: int):
    att = get_object_or_404(
        ExpenseAttachment.objects.select_related("document", "document__created_by"),
        pk=pk,
    )
    d = att.document
    if getattr(d, "doc_type", None) != "expense":
        messages.error(request, "Atașament invalid.")
        return redirect("ui:dashboard")

    if not can_modify_expense(request.user, d):
        messages.error(request, "Nu ai drepturi să ștergi acest atașament.")
        return redirect("ui:expense_detail", pk=d.id)

    next_url = safe_next_url(request, reverse("ui:expense_edit", args=[d.id]))

    if request.method == "POST":
        try:
            # ștergem și fișierul din storage
            if att.file:
                att.file.delete(save=False)
        except Exception:
            pass
        att.delete()
        messages.success(request, "Atașamentul a fost șters.")
        return redirect(next_url)

    return render(request, "ui/expense_attachment_confirm_delete.html", {"att": att, "d": d, "next_url": next_url})


@login_required
def sales_export_csv(request):
    """Export CSV pentru vânzări.

    Filtre (GET):
      - sales_from=YYYY-MM-DD
      - sales_to=YYYY-MM-DD
      - sales_buyer=string
      - sales_flock=<id>
      - sales_only_debts=1

    Coloane: DATA, ORA, CUMPĂRĂTOR, PUI ALBI, PUI COLORATI, FURAJ(KG), BANI, DATORIE
    """
    sales_from = parse_date((request.GET.get("sales_from") or "").strip())
    sales_to = parse_date((request.GET.get("sales_to") or "").strip())
    sales_buyer = (request.GET.get("sales_buyer") or "").strip()
    sales_flock = (request.GET.get("sales_flock") or "").strip()
    sales_only_debts = (request.GET.get("sales_only_debts") or "").strip() in ("1", "true", "True", "yes")
    sales_include_orphans = (request.GET.get("sales_include_orphans") or "").strip() in ("1", "true", "True", "yes")

    sales_qs = (
        Document.objects.filter(doc_type="sale")
        .select_related("flock", "flock__season", "flock__house", "partner")
        .prefetch_related("lines", "payments")
        .order_by("-date", "-id")
    )
    if not sales_include_orphans:
        sales_qs = sales_qs.filter(payments__isnull=False).distinct()
    if sales_from:
        sales_qs = sales_qs.filter(date__gte=sales_from)
    if sales_to:
        sales_qs = sales_qs.filter(date__lte=sales_to)
    if sales_buyer:
        sales_qs = sales_qs.filter(partner__name__icontains=sales_buyer)
    if sales_flock.isdigit():
        sales_qs = sales_qs.filter(flock_id=int(sales_flock))
    if sales_only_debts:
        sales_qs = sales_qs.filter(payments__status="due").distinct()

    filename = f"vanzari_{timezone.localdate().isoformat()}.csv"
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'

    # Excel-friendly BOM
    resp.write("\ufeff")

    writer = csv.writer(resp, delimiter=";")
    writer.writerow(["DATA", "ORA", "CUMPĂRĂTOR", "PUI ALBI", "PUI COLORATI", "FURAJ(KG)", "BANI", "DATORIE", "SERIE", "HALA"])

    def _sum_qty(doc: Document, *, keys: set[str]) -> Decimal:
        total = Decimal("0")
        for ln in doc.lines.all():
            desc = (ln.description or "").strip().lower()
            if desc in keys:
                total += (ln.qty or Decimal("0"))
        return total

    for d in sales_qs:
        pui_albi = int(_sum_qty(d, keys={"pui albi"}) or 0)
        pui_colorati = int(_sum_qty(d, keys={"pui colorați", "pui colorati"}) or 0)
        furaj = (_sum_qty(d, keys={"furaj"}) or Decimal("0")).quantize(Decimal("0.001"))
        datorie = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
        ora = timezone.localtime(d.created_at).strftime("%H:%M") if d.created_at else ""
        buyer = d.partner.name if d.partner else ""
        serie = d.flock.season.name if d.flock_id else ""
        hala = d.flock.house.name if d.flock_id else ""

        writer.writerow([
            d.date.isoformat() if d.date else "",
            ora,
            buyer,
            pui_albi,
            pui_colorati,
            str(furaj),
            str((d.total or Decimal("0.00")).quantize(Decimal("0.01"))),
            str(datorie),
            serie,
            hala,
        ])

    return resp

@login_required
def sales_export_xlsx(request):
    """Export Excel (.xlsx) pentru vânzări (aceleași filtre ca CSV)."""
    sales_from = parse_date((request.GET.get("sales_from") or "").strip())
    sales_to = parse_date((request.GET.get("sales_to") or "").strip())
    sales_buyer = (request.GET.get("sales_buyer") or "").strip()
    sales_flock = (request.GET.get("sales_flock") or "").strip()
    sales_only_debts = (request.GET.get("sales_only_debts") or "").strip() in ("1", "true", "True", "yes")
    sales_include_orphans = (request.GET.get("sales_include_orphans") or "").strip() in ("1", "true", "True", "yes")

    sales_qs = (
        Document.objects.filter(doc_type="sale")
        .select_related("flock", "flock__season", "flock__house", "partner")
        .prefetch_related("lines", "payments")
        .order_by("-date", "-id")
    )
    if not sales_include_orphans:
        sales_qs = sales_qs.filter(payments__isnull=False).distinct()
    if sales_from:
        sales_qs = sales_qs.filter(date__gte=sales_from)
    if sales_to:
        sales_qs = sales_qs.filter(date__lte=sales_to)
    if sales_buyer:
        sales_qs = sales_qs.filter(partner__name__icontains=sales_buyer)
    if sales_flock.isdigit():
        sales_qs = sales_qs.filter(flock_id=int(sales_flock))
    if sales_only_debts:
        sales_qs = sales_qs.filter(payments__status="due").distinct()

    def _sum_qty(doc: Document, *, keys: set[str]) -> Decimal:
        total = Decimal("0")
        for ln in doc.lines.all():
            desc = (ln.description or "").strip().lower()
            if desc in keys:
                total += (ln.qty or Decimal("0"))
        return total

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vanzari"

    headers = ["DATA", "ORA", "CUMPĂRĂTOR", "PUI ALBI", "PUI COLORATI", "FURAJ(KG)", "BANI", "DATORIE", "SERIE", "HALA"]
    ws.append(headers)

    for d in sales_qs:
        pui_albi = int(_sum_qty(d, keys={"pui albi"}) or 0)
        pui_colorati = int(_sum_qty(d, keys={"pui colorați", "pui colorati"}) or 0)
        furaj = (_sum_qty(d, keys={"furaj"}) or Decimal("0")).quantize(Decimal("0.001"))
        datorie = sum((p.amount for p in d.payments.all() if p.status == "due"), Decimal("0.00")).quantize(Decimal("0.01"))
        ora = timezone.localtime(d.created_at).strftime("%H:%M") if getattr(d, "created_at", None) else ""
        buyer = d.partner.name if d.partner else ""
        serie = d.flock.season.name if d.flock_id else ""
        hala = d.flock.house.name if d.flock_id else ""
        bani = sum((p.amount for p in d.payments.all() if p.status == "paid"), Decimal("0.00")).quantize(Decimal("0.01"))

        ws.append([
            d.date.isoformat() if d.date else "",
            ora,
            buyer,
            pui_albi,
            pui_colorati,
            float(furaj),
            float(bani),
            float(datorie),
            serie,
            hala,
        ])

    # formatare numerică simplă
    for row in ws.iter_rows(min_row=2):
        # FURAJ(KG) col 6
        row[5].number_format = "0.000"
        # BANI col 7, DATORIE col 8
        row[6].number_format = "0.00"
        row[7].number_format = "0.00"

    # lățimi coloane
    widths = [12, 8, 28, 10, 12, 12, 12, 12, 16, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"vanzari_{timezone.localdate().isoformat()}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



@login_required
def payment_mark_paid(request, pk: int):
    """Marchează o datorie (Payment status=due) ca fiind plătită."""
    p = get_object_or_404(
        Payment.objects.select_related("document", "document__partner", "created_by"),
        pk=pk,
    )

    next_url = (request.GET.get("next") or "").strip() or f"{reverse('ui:dashboard')}?tab=sales"

    if not can_modify_payment(request.user, p):
        messages.error(request, "Nu ai drepturi să modifici această plată.")
        return redirect(next_url)

    if p.status != "due":
        messages.info(request, "Această înregistrare nu mai este scadentă.")
        return redirect(next_url)

    if request.method == "POST":
        before = snapshot(p)
        paid_date = parse_date((request.POST.get("paid_date") or "").strip()) or timezone.localdate()
        method = (request.POST.get("method") or "cash").strip() or "cash"
        p.status = "paid"
        p.paid_date = paid_date
        p.method = method
        p.save(update_fields=["status", "paid_date", "method"])

        log_event(
            actor=request.user,
            action="UPDATE",
            instance=p,
            message=f"PAYMENT paid: #{p.id} ({p.amount} {p.document.currency})",
            before=before,
            after=snapshot(p),
            request=request,
        )
        messages.success(request, "Datoria a fost marcată ca plătită.")
        return redirect(next_url)

    # fallback GET
    return redirect(next_url)
